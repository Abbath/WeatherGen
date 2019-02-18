from openpyxl import load_workbook
import sys

header = """TERREL.DAT      2.0             Header structure with coordinate parameters                     
   1
Produced by TERREL Version: 3.3  Level: 030402                                  
UTM     
  34N   
WGS-G   10-10-2002  
      40      40     340.000    5504.000       1.000       1.000
KM  M   
W_E N_S
"""

header2 = """GEO.DAT         2.0             Header structure with coordinate parameters                     
   2
Produced by MAKEGEO Version: 2.2  Level: 030402                                 
Demo Application                                                                
UTM     
  34N   
WGS-G   10-10-2002  
      40      40     340.000    5504.000       1.000       1.000
KM  M   
 1          - LAND USE DATA - (1=new categories)
   14  50  55 - NLU, IWAT1, IWAT2
   10  20 -20  30  40  50  54  55  60  61  62  70  80  90
"""

def main(args):
    fname = args[1]
    res = header
    wb = load_workbook(fname)
    ws = wb.active
    m = 61
    n = 41
    counter = 0
    for row in ws.rows:
        v = row[1].value
        if isinstance(v, float):
            res += '{:8.1f}'.format(v)
            counter += 1
            if counter == 61:
                counter = 0
                res += '\n'
    fname0 = args[2] if len(args) > 2 else "res.dat"
    f = open(fname0, 'w')
    f.write(res)
    f.close()

def main2(args):
    fname = args[1]
    res = header2
    wb = load_workbook(fname, data_only=True)
    ws = wb.active
    m = 61
    n = 41
    w = 10
    counter = 0
    table = [[],[],[],[],[],[],[],[]]
    for row in ws.rows:
        if counter == 0:
            counter += 1
        else:
            counter += 1
            table[0].append(row[5].value)
            table[1].append(row[6].value)
            table[2].append(row[7].value)
            table[3].append(row[8].value)
            table[4].append(row[9].value)
            table[5].append(row[10].value)
            table[6].append(row[11].value)
            table[7].append(row[12].value)
    counter = 0
    counter1 = 0    
    for x in table[0]:
        if counter1 == len(table[0])-1:
            res += '{:7}'.format(x)
        else:
            res += '{:7},'.format(x)
        counter1 += 1
        counter += 1
        if counter == 10:
            counter = 0
            res += '\n'
    counter = 0
    counter1 = 0
    res += " 1.0000  - TERRAIN heights - HTFAC (Conversion to meters)\n"
    for x in table[1]:
        if counter1 == len(table[1])-1 or counter == 60:
            res += '{:7.2f}'.format(x)
        else:
            res += '{:7.2f},'.format(x)
        counter1 += 1
        counter += 1
        if counter == 61:
            counter = 0
            res += '\n'
    counter = 0
    counter1 = 0
    res += " 2    - gridded z0 field\n"
    for x in table[2]:
        if counter1 == len(table[2])-1 or counter == 60:
            res += '{:7.4f}'.format(x)
        else:
            res += '{:7.4f},'.format(x)
        counter1 += 1
        counter += 1
        if counter == 61:
            counter = 0
            res += '\n'
    counter = 0
    counter1 = 0
    res += " 2    - gridded albedo field\n"
    for x in table[3]:
        if counter1 == len(table[3])-1 or counter == 60:
            res += '{:7.2f}'.format(x)
        else:
            res += '{:7.2f},'.format(x)
        counter1 += 1
        counter += 1
        if counter == 61:
            counter = 0
            res += '\n'
    counter = 0
    counter1 = 0
    res += " 2    - gridded Bowen ratio field\n"
    for x in table[4]:
        if counter1 == len(table[4])-1 or counter == 60:
            res += '{:7.2f}'.format(x)
        else:
            res += '{:7.2f},'.format(x)
        counter1 += 1
        counter += 1
        if counter == 61:
            counter = 0
            res += '\n'
    counter = 0
    counter1 = 0
    res += " 2    - gridded soil heat flux parameters\n"
    for x in table[5]:
        if counter1 == len(table[5])-1 or counter == 60:
            res += '{:7.2f}'.format(x)
        else:
            res += '{:7.2f},'.format(x)
        counter1 += 1
        counter += 1
        if counter == 61:
            counter = 0
            res += '\n'
    counter = 0
    counter1 = 0
    res += " 2    - gridded anthropogenic heat flux field\n"
    for x in table[6]:
        if counter1 == len(table[6])-1 or counter == 60:
            res += '{:7.2f}'.format(x)
        else:
            res += '{:7.2f},'.format(x)
        counter1 += 1
        counter += 1
        if counter == 61:
            counter = 0
            res += '\n'
    counter = 0
    counter1 = 0
    res += " 2    - gridded leaf area index field\n"
    for x in table[7]:
        if counter1 == len(table[7])-1 or counter == 60:
            res += '{:7.2f}'.format(x)
        else:
            res += '{:7.2f},'.format(x)
        counter1 += 1
        counter += 1
        if counter == 61:
            counter = 0
            res += '\n'
    
    fname0 = args[2] if len(args) > 2 else "res.dat"
    f = open(fname0, 'w')
    f.write(res)
    f.close()

if __name__ == '__main__':
    if int(sys.argv[1]) == 1:
        main(sys.argv[1:])
    elif int(sys.argv[1]) == 2:
        main2(sys.argv[1:])